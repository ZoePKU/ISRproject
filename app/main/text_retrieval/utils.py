import jieba
import json
import openpyxl
jieba.load_userdict("main/text_retrieval/my_dictionary.txt")


def create_thes(filename):
    # 读取同义词表
    thes = openpyxl.load_workbook(filename)['replace']
    thes_dict = {}
    for i in range(640):
        formal = thes.cell(row=i + 1, column=1).value
        informal = thes.cell(row=i + 1, column=2).value
        if informal:
            informal_list = informal.split(';')
            for word in informal_list:
                thes_dict[word] = formal
    thes_words = set(thes_dict.keys())
    print("完成读取同义词表")
    return thes_words, thes_dict


def create_stopword(filename):
    # 读取停用词表
    stop_words = set()
    with open(filename, 'r', encoding='UTF-8') as f:
        for line in f.readlines():
            stop_words.add(line.rstrip('\n'))
    stop_words.update([' ', '\n'])
    print("完成读取停用词表")
    return stop_words


def init_thes():
    thes_words, thes_dict = create_thes('main/text_retrieval/thesaurus.xlsx')
    stop_words = create_stopword('main/text_retrieval/stop_words.txt')
    return thes_words, thes_dict, stop_words


# 输入一句话，返回分词结果的list
def parse(sentence, thes_words, thes_dict, stop_words):
    cut_list = jieba.lcut(sentence)
    new_cut_list = []
    for i in range(len(cut_list)):
        word = cut_list[i]
        if word in thes_words:
            word = thes_dict[word]
        if word not in stop_words:
            new_cut_list.append(word)
    return new_cut_list


# 输出json
def output(filename, content):
    with open(filename, 'w', encoding="utf-8") as f:
        json.dump(content, f, indent=4, ensure_ascii=False)
        f.close()
        print("输出完毕")


# 读入json,返回dict
def json_input(filename):
    with open(filename, "r", encoding="utf-8") as f:
        load_dict = json.load(f)
        return load_dict
