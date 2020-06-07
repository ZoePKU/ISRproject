import openpyxl
'''
with open("main/cnn_retrieval/out.log","r",encoding = "utf-8") as f:
    content = f.read()
    lst = content.split("\n")
    print(lst)
    new_lst = [x.split() for x in lst]
    print(new_lst)
    # 读取图片描述
    des = openpyxl.load_workbook('main/text_retrieval/bqb_description.xlsx')['bqb_description']
    des_dict = {}
    for i in range(4000):
        id = des.cell(row=i + 1, column=1).value
        description = des.cell(row=i + 1, column=2).value
        des_dict[id] = description
    print(des_dict)
    print("完成读取图片描述")

    for x in new_lst:
        print(x[1],x[2],des_dict["{:0>4}".format(x[1])],des_dict["{:0>4}".format(x[2])])
'''
with open("main/cnn_retrieval/out_2.log","r",encoding = "utf-8") as f:
    content = f.read()
    lst = content.split("\n")
    print(lst)
    new_lst = [x.split() for x in lst]
    print(new_lst)
    my_set = set()
    length = len(new_lst)
    for i in range(length):
        my_set.add("{:0>4}".format(new_lst[i][2]))
    print(len(my_set))
    lst = sorted(list(my_set))
    print(lst)