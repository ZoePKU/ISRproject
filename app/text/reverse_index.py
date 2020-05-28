import jieba
import openpyxl
from collections import Counter
import json
from text.utils import *


def description_parse(dict, thes_words, thes_dict):
    cut_dict = {}
    for id in dict:
        if dict[id]:
            cut_dict[id] = parse(dict[id], thes_words, thes_dict)
        else:
            cut_dict[id] = ""
    return cut_dict


def revert(cut_dict):
    reverse_index = {}
    all_words = Counter()
    words_num = 0
    for id in cut_dict:
        words_num += len(cut_dict[id])
        all_words.update(cut_dict[id])
    set_all_words = set(all_words.elements())
    for word in set_all_words:
        temp = {}
        tf = all_words[word] / words_num
        idf = 1  # 待 定
        for id in cut_dict:
            if word in cut_dict[id]:
                temp[id] = tf * idf
        reverse_index[word] = temp
    return reverse_index


if __name__ == '__main__':

    # 生成词表
    thes_words, stop_words = init_thes()

    # 读取图片描述
    des = openpyxl.load_workbook('text/bqb_description.xlsx')['bqb_description']
    des_dict = {}
    for i in range(4000):
        id = des.cell(row=i + 1, column=1).value
        description = des.cell(row=i + 1, column=2).value
        des_dict[id] = description
    print("完成读取图片描述")

    # 分词
    cut_dict = description_parse(des_dict, thes_words, stop_words)
    print("完成分词")

    # 建倒排索引
    reverse_index = revert(cut_dict)
    print("完成建立倒排档")

    # index_json = json.dumps(reverse_index, sort_keys=True, ensure_ascii=False, indent=4)
    # 输出到文件
    output('text/reverse_index.json', reverse_index)
    load_dict = input('text/reverse_index.json')
    # print(load_dict)
